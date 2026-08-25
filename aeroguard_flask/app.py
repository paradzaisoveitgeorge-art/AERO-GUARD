"""AERO-GUARD Flask MVP — provider console + Smartpoint consultant demo.

Section 2 of the MVP build introduces a real SQLite database (via
SQLAlchemy + Flask-Migrate). Every entity that used to live in an
in-memory list (agencies, escalations, threads, etc.) now
persists to disk. Use the CLI commands at the bottom of this file to
seed and reset demo data:

    flask --app app seed          # populate the standard demo state
    flask --app app reset-demo    # wipe and re-seed

Multi-tenancy: every operational row carries a ``provider_id``. Until
auth lands in Section 3 we read all data globally so the visual demo
keeps working; the schema is already correct for tenant scoping later.
"""
from __future__ import annotations

import base64
import csv
import io
import json
import os
import random
import secrets
from pathlib import Path

from datetime import datetime, timedelta

from dotenv import load_dotenv
from flask import (
    Flask,
    Response,
    abort,
    flash,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_login import (
    LoginManager,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_migrate import Migrate
from flask_wtf.csrf import CSRFProtect
from flask import jsonify

from models import (
    Agency,
    AgencyMember,
    Alert,
    AuditLog,
    Broadcast,
    Escalation,
    LearningModule,
    Message,
    PendingIssue,
    PolicyDoc,
    Provider,
    Thread,
    TicketIssue,
    User,
    db,
)
from permissions import can, require, require_role
from mailer import send_email, is_configured as mail_is_configured

# Load .env if present (no-op in production where vars come from the host)
load_dotenv()

app = Flask(__name__)

# --- Configuration ----------------------------------------------------------
_env = os.environ.get("FLASK_ENV", "development").lower()
_is_prod = _env == "production"

_secret = os.environ.get("SECRET_KEY")
if not _secret:
    if _is_prod:
        raise RuntimeError("SECRET_KEY env var is required when FLASK_ENV=production")
    _secret = secrets.token_hex(32)
app.config["SECRET_KEY"] = _secret
app.config["DEBUG"] = not _is_prod

# DATABASE_URL — Postgres on production hosts, SQLite locally.
_default_db = "sqlite:///" + str(Path(app.instance_path) / "aeroguard.db")
db_url = os.environ.get("DATABASE_URL", _default_db)
# Render/Heroku give us postgres:// but SQLAlchemy wants postgresql://
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Session cookie hardening. Secure cookies are only sent over HTTPS; we
# disable that locally so dev login works on http://localhost.
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = _is_prod
app.config["REMEMBER_COOKIE_HTTPONLY"] = True
app.config["REMEMBER_COOKIE_SAMESITE"] = "Lax"
app.config["REMEMBER_COOKIE_SECURE"] = _is_prod
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=12)
# CSRF tokens never expire mid-session (avoids "form too old" errors on
# long demo sessions). We still rotate them each login.
app.config["WTF_CSRF_TIME_LIMIT"] = None

Path(app.instance_path).mkdir(parents=True, exist_ok=True)

db.init_app(app)
Migrate(app, db)
csrf = CSRFProtect(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Please sign in to continue."
login_manager.login_message_category = "info"


@login_manager.user_loader
def _load_user(user_id: str):
    return User.query.get(user_id)


# Login rate limiter — 5 attempts/min per IP. Runs in-memory; a real
# deploy would back this with Redis. Fine for the MVP.
limiter = Limiter(get_remote_address, app=app, default_limits=[])


@app.context_processor
def inject_user():
    """Make ``current_user`` + permission helpers available in every template."""
    provider = None
    alerts_count = 0
    chat_unread = 0
    if current_user.is_authenticated and current_user.provider_id:
        provider = Provider.query.get(current_user.provider_id)
        alerts_count = Escalation.query.filter_by(
            provider_id=current_user.provider_id, status="OPEN"
        ).count()
        chat_unread = (
            db.session.query(db.func.coalesce(db.func.sum(Thread.unread), 0))
            .filter(Thread.provider_id == current_user.provider_id)
            .scalar()
        ) or 0
    return {
        "current_user": current_user,
        "current_provider": provider,
        "topbar_alerts_count": alerts_count,
        "topbar_chat_unread": int(chat_unread),
        "can": lambda action: can(current_user, action),
        "humanize": humanize,
        "humanize_sla": humanize_sla,
    }


# Public endpoints that never require login. Everything else under
# /provider/* and / is locked down by the before_request guard below.
PUBLIC_ENDPOINTS = {"login", "login_mfa", "login_mfa_cancel", "logout", "forgot", "reset_password", "healthz", "legal", "static",
                    "api_compliance_validate", "api_compliance_escalate", "api_ticket_status"}


@app.errorhandler(403)
def _forbidden(_e):
    return render_template("auth/403.html"), 403


@app.errorhandler(404)
def _not_found(_e):
    return render_template("auth/404.html"), 404


@app.errorhandler(500)
def _server_error(_e):
    return render_template("auth/500.html"), 500


@app.before_request
def _require_login_for_provider_console():
    if request.endpoint in PUBLIC_ENDPOINTS or request.endpoint is None:
        return
    if not current_user.is_authenticated:
        return redirect(url_for("login", next=request.path))
    # Consultants and agency users cannot access /provider/* (the console).
    if request.path.startswith("/provider") and not current_user.is_provider_staff():
        abort(403)
    # The Agency Portal is only for agency-portal roles.
    if request.path.startswith("/portal") and not current_user.is_agency_user():
        abort(403)


# --- Static reference data (still in code — never changes per-tenant) -------
COUNTRIES = ["ZW", "ZA", "KE", "UG", "TZ", "NG", "GH", "ET", "RW", "BW"]
# Subscription tiers per the client spec (Platinum / Gold / Standard).
POLICY_TEMPLATES = ["Platinum", "Gold", "Standard", "Custom"]
REGIONS = ["Southern Africa", "East Africa", "West Africa", "Central Africa", "North Africa", "Indian Ocean"]
CURRENCIES = ["USD", "EUR", "ZAR", "KES", "UGX", "NGN", "GHS", "ZWL", "RWF"]

NAV_GROUPS = [
    {"label": "OVERVIEW", "items": [
        {"key": "DASHBOARD", "label": "Dashboard", "icon": "▦", "endpoint": "provider_dashboard"},
    ]},
    {"label": "OPERATIONAL CONTROL", "items": [
        {"key": "AGENCIES", "label": "Agency Provisioning", "icon": "\U0001f3e2", "endpoint": "provider_agencies"},
        {"key": "USERS", "label": "Helpdesk Users", "icon": "\U0001f465", "endpoint": "provider_users"},
    ]},
    {"label": "INTELLIGENCE", "items": [
        {"key": "AUDITS", "label": "Agency ADM Audits", "icon": "\U0001f4ca", "endpoint": "provider_audits"},
        {"key": "AUDIT_LOG", "label": "Audit Log", "icon": "\U0001f4dd", "endpoint": "provider_audit_log"},
    ]},
    {"label": "SUPPORT TOOLS", "items": [
        {"key": "ESCALATIONS", "label": "Escalations", "icon": "⚠", "endpoint": "provider_escalations"},
        {"key": "EMULATE", "label": "Emulate into PCC", "icon": "⌨", "endpoint": "provider_emulate"},
        {"key": "RESPOND", "label": "Respond to Clients", "icon": "\U0001f4ac", "endpoint": "provider_respond"},
    ]},
    {"label": "KNOWLEDGE", "items": [
        {"key": "LEARNING", "label": "My Learning", "icon": "\U0001f393", "endpoint": "provider_learning"},
        {"key": "POLICIES", "label": "Terms & Policies", "icon": "\U0001f4dc", "endpoint": "provider_policies"},
    ]},
]

ROLE_HINTS = {
    "L1": "Ticket support, view dashboards, respond to clients",
    "L2": "Above + Emulate PCC, escalate to vendors",
    "ADMIN": "Full control: provision agencies, manage users, audit trail",
}

REASON_DIST = [
    {"label": "Duplicate booking",     "pct": 32, "color": "bar-rose"},
    {"label": "Fare rule violation",   "pct": 24, "color": "bar-amber"},
    {"label": "Time limit expiry",     "pct": 18, "color": "bar-indigo"},
    {"label": "Schedule change",       "pct": 14, "color": "bar-sky"},
    {"label": "Other",                 "pct": 12, "color": "bar-slate"},
]

DRILLDOWN_RULES = [
    {"code": "FXR-103", "tone": "amber",  "text": "Fare basis mismatch · 4 PNRs · est $1,600"},
    {"code": "TKT-204", "tone": "rose",   "text": "Ticketing time limit expired · 2 PNRs · est $800"},
    {"code": "NCP-011", "tone": "indigo", "text": "Name change post-ticketing · 1 PNR · est $320"},
]

TUTORIALS = [
    {"id": "passport-scan", "title": "Passport Auto-Fill & MRZ Scan",  "blurb": "Drop a passport image — AERO-GUARD reads the MRZ, validates ICAO 9303, and pushes DOCS SSR to the PNR. Zero spelling errors.", "duration": "1:42", "tag": "DOCS · OCR"},
    {"id": "pnr-validator", "title": "Live PNR Rule Validator",        "blurb": "Watch AERO-GUARD intercept a min-stay breach mid-pricing and suggest the compliant fare basis before ticketing.",         "duration": "2:15", "tag": "ADM · Rules"},
    {"id": "adm-watch",     "title": "ADM Watch & Audit Trail",        "blurb": "End-to-end demo: catch a tax-code violation, resolve it, and audit the trail from the helpdesk console.",                 "duration": "2:58", "tag": "ADM · Audit"},
]


# --- Time helpers ----------------------------------------------------------

def humanize(when, *, future_ok: bool = False) -> str:
    """Convert a datetime → '3 min ago', 'just now', 'in 2 hr', '—' if None."""
    if when is None:
        return "—"
    now = datetime.utcnow()
    delta = now - when
    seconds = delta.total_seconds()
    if seconds < 0:
        if not future_ok:
            return "just now"
        seconds = -seconds
        suffix = lambda s: f"in {s}"
    else:
        suffix = lambda s: f"{s} ago"
    if seconds < 45:
        return "just now" if seconds >= 0 and delta >= timedelta(0) else suffix("a few seconds")
    minutes = seconds / 60
    if minutes < 60:
        return suffix(f"{int(minutes)} min")
    hours = minutes / 60
    if hours < 24:
        return suffix(f"{int(hours)} hr")
    days = hours / 24
    if days < 2:
        return "yesterday" if not future_ok else "tomorrow"
    if days < 30:
        return suffix(f"{int(days)} days")
    if days < 365:
        return suffix(f"{int(days/30)} mo")
    return suffix(f"{int(days/365)} yr")


def humanize_sla(due: "datetime | None") -> str:
    """Render SLA deadline: 'X left' if future, 'overdue Yh' if past."""
    if due is None:
        return "—"
    now = datetime.utcnow()
    delta = due - now
    secs = delta.total_seconds()
    if secs <= 0:
        # past due
        secs = -secs
        if secs < 3600:
            return f"overdue {int(secs / 60)} min"
        if secs < 86400:
            return f"overdue {int(secs / 3600)} hr"
        return f"overdue {int(secs / 86400)} days"
    if secs < 3600:
        return f"{int(secs / 60)} min left"
    if secs < 86400:
        return f"{int(secs / 3600)} hr left"
    return f"{int(secs / 86400)} days left"


# --- Tenancy helpers -------------------------------------------------------

def current_provider_id() -> str | None:
    """The provider_id of the logged-in user, or None for consultants/anon."""
    return getattr(current_user, "provider_id", None)


def tenant_q(model):
    """Return ``model.query`` filtered to the current user's provider.

    Use for every list/search endpoint that touches a tenant-scoped table
    (Agency, User, Escalation, Thread). Catalog tables
    (PolicyDoc, LearningModule, Alert, PendingIssue) are shared and
    should query the model directly.
    """
    return model.query.filter_by(provider_id=current_provider_id())


def get_owned_or_404(model, pk):
    """Single-row lookup that 404s if the row is missing OR not ours.

    We return 404 (not 403) on cross-tenant access so the response is
    indistinguishable from "doesn't exist" — that way an attacker can't
    use the error code to probe which IDs belong to other tenants.
    """
    obj = model.query.get(pk)
    if obj is None:
        abort(404)
    if getattr(obj, "provider_id", None) != current_provider_id():
        abort(404)
    return obj


def write_audit(action: str, target_type: str, target_id: str, *, note: str = "") -> None:
    """Append a row to the audit log. Always tagged with provider + actor."""
    db.session.add(AuditLog(
        provider_id=current_provider_id(),
        actor_user_id=getattr(current_user, "id", None),
        action=action,
        target_type=target_type,
        target_id=str(target_id),
        note=note or None,
    ))


# --- Helpers ---------------------------------------------------------------

def render_provider(template_name, active_nav, **context):
    groups = [
        {
            "label": g["label"],
            "items": [
                {**item, "url": url_for(item["endpoint"])} for item in g["items"]
            ],
        }
        for g in NAV_GROUPS
    ]
    return render_template(template_name, nav_groups=groups, active_nav=active_nav, **context)


def policy_tier(policy_level: str) -> str:
    """Map the free-text policy_level onto the spec's service-tier naming
    (Platinum / Gold / Standard). Keyword-based so it copes with the new
    subscription-tier vocabulary (Platinum / Gold / Standard) as well as
    legacy values still in older databases (STANDARD / ENTERPRISE / BASIC,
    Full Enterprise / Standard Compliance / Trial / Custom)."""
    p = (policy_level or "").upper()
    if "ENTERPRISE" in p or "PLATINUM" in p:
        return "PLATINUM"
    if "GOLD" in p:
        return "GOLD"
    if "BASIC" in p or "TRIAL" in p or "LITE" in p:
        return "STANDARD"
    if "CUSTOM" in p:
        return "CUSTOM"
    if "STANDARD" in p:
        return "STANDARD"
    return "GOLD"


def agency_to_dict(a: Agency) -> dict:
    """Templates were written against dicts — keep that contract."""
    return {
        "id": a.id, "name": a.name, "pcc": a.pcc, "gds": a.gds, "country": a.country,
        "seats": a.seats, "used_seats": a.used_seats, "status": a.status,
        "month_adms": a.month_adms,
        "last_active": humanize(a.updated_at) if a.status != "PROVISIONING" else "—",
        "policy_level": a.policy_level, "policy_tier": policy_tier(a.policy_level),
        "admin_email": a.admin_email, "region": a.region or "",
    }


ESCALATION_TIERS = {
    "GENERAL": "Tier 1 · General",
    "FINANCIAL": "Tier 2 · Financial",
    "TECHNICAL": "Tier 3 · Technical",
}


def escalation_to_dict(e: Escalation) -> dict:
    category = e.category or "GENERAL"
    return {
        "id": e.id, "agency": e.agency, "pnr": e.pnr, "subject": e.subject,
        "level": e.level, "priority": e.priority,
        "category": category,
        "tier_label": ESCALATION_TIERS.get(category, ESCALATION_TIERS["GENERAL"]),
        "opened": humanize(e.created_at),
        "status": e.status,
        "sla": humanize_sla(e.sla_due_at),
    }


def thread_to_dict(t: Thread) -> dict:
    return {
        "id": t.id, "agency": t.agency, "agent": t.agent, "unread": t.unread, "last": t.last,
        "messages": [{"from": m.sender, "text": m.text, "t": m.t} for m in t.messages],
    }


def all_agencies() -> list[Agency]:
    """Tenant-scoped, non-deleted agency list for list endpoints + dropdowns."""
    return (
        tenant_q(Agency)
        .filter(Agency.deleted_at.is_(None))
        .order_by(Agency.created_at)
        .all()
    )


def _landing_url_for(user: User) -> str:
    """Where a user lands after login: provider staff → console,
    agency users → their portal, consultants → smartpoint."""
    if user.is_provider_staff():
        return url_for("provider_dashboard")
    if user.is_agency_user():
        return url_for("portal_dashboard")
    return url_for("smartpoint_demo")


# --- Routes: auth ---------------------------------------------------------

def _finalize_login(user, remember: bool, next_url: str):
    """Complete a successful login and route the user to their landing page."""
    login_user(user, remember=remember)
    user.last_login_at = datetime.utcnow()
    user.last_login = "now"  # legacy display string kept in sync
    db.session.commit()
    session.pop("mfa_pending_user_id", None)
    session.pop("mfa_pending_remember", None)
    session.pop("mfa_pending_next", None)
    # Only honour next_url if it's a local path (open-redirect guard).
    if next_url and next_url.startswith("/") and not next_url.startswith("//"):
        return redirect(next_url)
    return redirect(_landing_url_for(user))


@app.route("/login", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def login():
    if current_user.is_authenticated:
        return redirect(_landing_url_for(current_user))

    error = None
    next_url = request.args.get("next") or request.form.get("next") or ""

    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        remember = bool(request.form.get("remember"))

        user = User.query.filter(db.func.lower(User.email) == email).first()
        if user is None or not user.check_password(password):
            error = "Invalid email or password."
        elif not user.active:
            error = "This account is disabled. Contact your administrator."
        elif user.mfa_secret:
            # Password OK, but MFA is enrolled — stash context and challenge.
            session["mfa_pending_user_id"] = user.id
            session["mfa_pending_remember"] = remember
            session["mfa_pending_next"] = next_url
            return redirect(url_for("login_mfa"))
        else:
            return _finalize_login(user, remember, next_url)

    return render_template("auth/login.html", error=error, next=next_url)


@app.route("/login/mfa", methods=["GET", "POST"])
@limiter.limit("10 per minute", methods=["POST"])
def login_mfa():
    """Second-factor challenge after a successful password step.

    Accepts either the live TOTP code from the authenticator app OR one
    of the user's single-use backup codes — the latter is consumed on
    use so a stolen sheet has a short shelf life.
    """
    import pyotp

    pending_id = session.get("mfa_pending_user_id")
    if not pending_id:
        return redirect(url_for("login"))
    user = User.query.get(pending_id)
    if user is None or not user.mfa_secret or not user.active:
        session.pop("mfa_pending_user_id", None)
        return redirect(url_for("login"))

    error = None
    if request.method == "POST":
        raw = (request.form.get("code") or "").strip()
        totp_code = raw.replace(" ", "")
        totp = pyotp.TOTP(user.mfa_secret)
        # `valid_window=1` accepts the previous 30s step too (clock skew).
        totp_ok = totp.verify(totp_code, valid_window=1)
        backup_ok = False
        if not totp_ok:
            backup_ok = user.consume_backup_code(raw)
        if totp_ok or backup_ok:
            if backup_ok:
                # Persist the consumed code + note it in the audit log
                # so a real user can trace unexpected uses later.
                db.session.commit()
                write_audit(
                    "MFA_BACKUP_CODE_USED", "user", user.id,
                    note=f"remaining={user.mfa_backup_codes_remaining}",
                )
                db.session.commit()
                # Warn them if they're running low.
                if user.mfa_backup_codes_remaining <= 2:
                    flash(
                        f"You have {user.mfa_backup_codes_remaining} backup "
                        "code(s) left. Generate a fresh set after signing in.",
                        "info",
                    )
            remember = bool(session.get("mfa_pending_remember"))
            next_url = session.get("mfa_pending_next") or ""
            return _finalize_login(user, remember, next_url)
        error = "That code is not valid. Try again — codes rotate every 30 seconds."

    return render_template("auth/mfa_challenge.html", error=error, email=user.email)


@app.route("/login/mfa/cancel", methods=["POST"])
def login_mfa_cancel():
    """Abandon the MFA challenge and return to the password screen."""
    session.pop("mfa_pending_user_id", None)
    session.pop("mfa_pending_remember", None)
    session.pop("mfa_pending_next", None)
    return redirect(url_for("login"))


def _generate_backup_codes(n: int = 10) -> list[str]:
    """Return `n` fresh 8-char uppercase-hex recovery codes.

    Displayed to the user hyphenated (e.g. `A3F1-B7C2`) for readability;
    matching in `User.consume_backup_code` strips the hyphen.
    """
    import secrets as _secrets
    return [_secrets.token_hex(4).upper() for _ in range(n)]


@app.route("/account/mfa/enroll", methods=["GET", "POST"])
@login_required
def mfa_enroll():
    """Show a QR + secret for the user to scan, then verify a code to activate."""
    import pyotp
    import qrcode

    # Reuse an in-progress secret in the session so refreshing the page
    # doesn't reset the QR (which would break Authenticator setup).
    pending_secret = session.get("mfa_enroll_secret")
    if not pending_secret:
        pending_secret = pyotp.random_base32()
        session["mfa_enroll_secret"] = pending_secret

    error = None
    if request.method == "POST":
        code = (request.form.get("code") or "").strip().replace(" ", "")
        totp = pyotp.TOTP(pending_secret)
        if totp.verify(code, valid_window=1):
            current_user.mfa_secret = pending_secret
            current_user.mfa = True
            # Generate + persist backup codes here (only chance to
            # show them in plaintext — after this we only ever hold
            # the hashes).
            plain_codes = _generate_backup_codes()
            current_user.set_backup_codes(plain_codes)
            db.session.commit()
            write_audit("MFA_ENABLED", "user", current_user.id)
            db.session.commit()
            session.pop("mfa_enroll_secret", None)
            session["mfa_new_backup_codes"] = plain_codes
            flash("Multi-factor authentication is now enabled.", "success")
            return redirect(url_for("mfa_recovery_codes"))
        error = "That code is not valid. Try again — codes rotate every 30 seconds."

    # Build a provisioning URI for the authenticator app.
    provisioning_uri = pyotp.TOTP(pending_secret).provisioning_uri(
        name=current_user.email,
        issuer_name="AERO-GUARD",
    )
    # Render the QR code as an inline data-URI PNG so we don't need
    # a separate endpoint.
    img = qrcode.make(provisioning_uri)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    qr_data_uri = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")

    return render_template(
        "auth/mfa_enroll.html",
        qr_data_uri=qr_data_uri,
        secret=pending_secret,
        error=error,
    )


@app.route("/account/mfa/disable", methods=["POST"])
@login_required
def mfa_disable():
    """User disables their own MFA. Requires the current TOTP code."""
    import pyotp

    if not current_user.mfa_secret:
        flash("MFA isn't enabled on your account.", "info")
        return redirect(url_for("_landing_url_for_current"))

    code = (request.form.get("code") or "").strip().replace(" ", "")
    if not pyotp.TOTP(current_user.mfa_secret).verify(code, valid_window=1):
        flash("Wrong code — MFA was not disabled.", "error")
        return redirect(url_for("mfa_enroll"))

    current_user.mfa_secret = None
    current_user.mfa = False
    # Drop backup codes too — they're only useful when MFA is enabled.
    current_user.mfa_backup_codes = None
    db.session.commit()
    write_audit("MFA_DISABLED", "user", current_user.id)
    db.session.commit()
    flash("Multi-factor authentication is now disabled.", "info")
    return redirect(url_for("_landing_url_for_current"))


@app.route("/account/mfa/recovery-codes", methods=["GET", "POST"])
@login_required
def mfa_recovery_codes():
    """One-time display of freshly-generated backup codes.

    Codes come from the session (set on enrollment or regenerate) — we
    never store them plaintext. POST acknowledges the user saw them and
    drops them from the session.
    """
    codes = session.get("mfa_new_backup_codes")
    if request.method == "POST":
        session.pop("mfa_new_backup_codes", None)
        return redirect(url_for("_landing_url_for_current"))
    if not codes:
        # Nothing to reveal — send the user to the MFA page where
        # they can generate a new set.
        return redirect(url_for("mfa_enroll"))
    return render_template("auth/mfa_recovery_codes.html", codes=codes)


@app.route("/account/mfa/regenerate-codes", methods=["POST"])
@login_required
@limiter.limit("5 per hour")
def mfa_regenerate_codes():
    """Issue a fresh set of backup codes. Requires the current TOTP so a
    hijacked session can't silently rotate codes."""
    import pyotp

    if not current_user.mfa_secret:
        flash("MFA isn't enabled on your account.", "info")
        return redirect(url_for("_landing_url_for_current"))

    code = (request.form.get("code") or "").strip().replace(" ", "")
    if not pyotp.TOTP(current_user.mfa_secret).verify(code, valid_window=1):
        flash("Wrong code — recovery codes were not regenerated.", "error")
        return redirect(url_for("mfa_enroll"))

    plain_codes = _generate_backup_codes()
    current_user.set_backup_codes(plain_codes)
    db.session.commit()
    write_audit("MFA_RECOVERY_REGENERATED", "user", current_user.id)
    db.session.commit()
    session["mfa_new_backup_codes"] = plain_codes
    return redirect(url_for("mfa_recovery_codes"))


@app.route("/_go", endpoint="_landing_url_for_current")
@login_required
def _landing_url_for_current():
    """Tiny helper endpoint so url_for('_landing_url_for_current') works."""
    return redirect(_landing_url_for(current_user))


@app.route("/logout", methods=["GET", "POST"])
@login_required
def logout():
    # GET renders a confirm-sign-out page (so bookmarks / typed URLs
    # don't 405). Only POST actually clears the session — that keeps
    # logout CSRF-protected.
    if request.method == "GET":
        return render_template("auth/logout_confirm.html")
    logout_user()
    flash("Signed out.", "info")
    return redirect(url_for("login"))


@app.route("/forgot", methods=["GET", "POST"])
@limiter.limit("3 per minute", methods=["POST"])
def forgot():
    """Issue a reset token and email the link. In dev (no SMTP configured)
    the link is also shown on-screen so demos still work end-to-end."""
    import secrets as _secrets

    reset_link = None
    submitted = False
    email_sent = False
    if request.method == "POST":
        submitted = True
        email = (request.form.get("email") or "").strip().lower()
        user = User.query.filter(db.func.lower(User.email) == email).first()
        if user and user.active:
            token = _secrets.token_urlsafe(24)
            user.reset_token = token
            user.reset_expires = datetime.utcnow() + timedelta(hours=2)
            db.session.commit()
            link = url_for("reset_password", token=token, _external=True)
            email_sent = send_email(
                to=user.email,
                subject="Reset your AERO-GUARD password",
                text=(
                    f"Hi {user.name},\n\n"
                    "Someone requested a password reset for your AERO-GUARD account. "
                    "If that was you, follow the link below within the next 2 hours:\n\n"
                    f"{link}\n\n"
                    "If it wasn't you, ignore this message — your password stays unchanged.\n\n"
                    "— AERO-GUARD"
                ),
                html=(
                    f"<p>Hi {user.name},</p>"
                    "<p>Someone requested a password reset for your AERO-GUARD account. "
                    "If that was you, follow the link below within the next 2 hours:</p>"
                    f'<p><a href="{link}">{link}</a></p>'
                    "<p>If it wasn't you, ignore this message — your password stays unchanged.</p>"
                    "<p>&mdash; AERO-GUARD</p>"
                ),
            )
            # Only reveal the link on-page when we couldn't actually
            # send it. Prevents email-enumeration in production.
            if not email_sent:
                reset_link = link
    return render_template(
        "auth/forgot.html",
        submitted=submitted,
        reset_link=reset_link,
        email_sent=email_sent,
        mail_configured=mail_is_configured(),
    )


@app.route("/reset/<token>", methods=["GET", "POST"])
@limiter.limit("20 per minute", methods=["GET"])   # blunt brute-force of the token itself
@limiter.limit("5 per minute", methods=["POST"])   # weak-password retry cap on a known token
def reset_password(token: str):
    user = User.query.filter_by(reset_token=token).first()
    # A disabled account must not be able to reset its way back in.
    expired = (
        user is None
        or user.reset_expires is None
        or user.reset_expires < datetime.utcnow()
        or not user.active
    )
    error = None
    done = False
    if request.method == "POST":
        if expired:
            error = "Reset link is invalid or expired."
        else:
            pw = request.form.get("password") or ""
            if len(pw) < 8:
                error = "Password must be at least 8 characters."
            else:
                user.set_password(pw)
                user.reset_token = None
                user.reset_expires = None
                db.session.commit()
                done = True
                # Confirmation email so a silent account takeover
                # (attacker with a stolen link) surfaces to the real
                # owner instead of staying quiet. Best-effort — a mail
                # failure must not roll back the successful reset.
                send_email(
                    to=user.email,
                    subject="Your AERO-GUARD password was changed",
                    text=(
                        f"Hi {user.name},\n\n"
                        "Your AERO-GUARD password was just changed. If that was you, "
                        "no action is needed.\n\n"
                        "If you didn't do this, contact your provider admin right "
                        "away — someone else may have access to your inbox.\n\n"
                        "— AERO-GUARD"
                    ),
                    html=(
                        f"<p>Hi {user.name},</p>"
                        "<p>Your AERO-GUARD password was just changed. If that was you, "
                        "no action is needed.</p>"
                        "<p>If you didn't do this, contact your provider admin right "
                        "away &mdash; someone else may have access to your inbox.</p>"
                        "<p>&mdash; AERO-GUARD</p>"
                    ),
                )
    return render_template("auth/reset.html", expired=expired, error=error, done=done, token=token)


# --- Routes: consultant view ----------------------------------------------

@app.route("/")
@login_required
def smartpoint_demo():
    # Provider staff may open the module replica for troubleshooting
    # (Products → Smart Button replica); otherwise they go to the console.
    is_replica = request.args.get("replica") == "1"
    if current_user.is_provider_staff() and not is_replica:
        return redirect(url_for("provider_dashboard"))
    if current_user.is_agency_user():
        return redirect(url_for("portal_dashboard"))
    # The demo terminal is bound to Skylink Travel — its branding toggle
    # decides how the itinerary trust badge renders (SB-10).
    demo_agency = Agency.query.filter_by(name=DEMO_TERMINAL["agency"]).first()
    badge_style = (demo_agency.badge_style if demo_agency and demo_agency.badge_style else "PROMINENT")
    return render_template("smartpoint.html", tutorials=TUTORIALS,
                           badge_style=badge_style, is_replica=is_replica)


# The demo consultant terminal is bound to Skylink Travel under AERO-GUARD
# HQ — the tenant mapping a real GDS plugin would derive from its PCC.
DEMO_TERMINAL = {"provider_id": "PRV-AG", "agency": "Skylink Travel", "pcc": "HREOU"}


@app.route("/api/consultant/notify-helpdesk", methods=["POST"])
def consultant_notify_helpdesk():
    """#AG Queue-on-Demand: file a real escalation carrying the PNR context
    so it appears on the provider's Escalations queue instantly."""
    payload = request.get_json(silent=True) or {}
    pnr = (payload.get("pnr") or "—").upper()[:20]
    subject = (payload.get("subject") or "Consultant assistance requested")[:200]
    priority = "HIGH" if payload.get("severity") == "CRITICAL" else "MED"
    new_id = f"ESC-{random.randint(8000, 8999)}"
    e = Escalation(
        id=new_id,
        provider_id=DEMO_TERMINAL["provider_id"],
        agency=DEMO_TERMINAL["agency"],
        pnr=pnr,
        subject=f"[#AG] {subject}",
        level="L1",
        priority=priority,
        opened="just now",
        status="OPEN",
        sla="4 hr left",
    )
    e.sla_due_at = datetime.utcnow() + timedelta(hours=4)
    db.session.add(e)
    # write_audit() tags rows with the actor's provider, which a consultant
    # doesn't have — log against the terminal's tenant instead.
    db.session.add(AuditLog(
        provider_id=DEMO_TERMINAL["provider_id"],
        actor_user_id=getattr(current_user, "id", None),
        action="AG_NOTIFY_HELPDESK",
        target_type="escalation",
        target_id=new_id,
        note=f"PNR {pnr} · {subject}",
    ))
    db.session.commit()
    return jsonify({
        "status": "QUEUED",
        "ticket_id": new_id,
        "message": "Escalation logged. Helpdesk supervisor notified. Average review time: 4 minutes.",
    })


@app.route("/api/consultant/chat-context", methods=["POST"])
def consultant_chat_context():
    """#AG live-chat bridge: push the consultant's PNR context into the
    provider-side support thread (the Agency Portal chat lands in a later
    batch — the deep-link contract is already honoured here)."""
    payload = request.get_json(silent=True) or {}
    pnr = (payload.get("pnr") or "—").upper()[:20]
    text = (payload.get("message") or f"Chat opened from #AG · PNR {pnr}")[:255]
    t = (Thread.query
         .filter_by(provider_id=DEMO_TERMINAL["provider_id"], agency=DEMO_TERMINAL["agency"])
         .first())
    if t is None:
        t = Thread(id=f"T-{random.randint(100, 999)}",
                   provider_id=DEMO_TERMINAL["provider_id"],
                   agency=DEMO_TERMINAL["agency"],
                   agent=getattr(current_user, "name", "Agent"),
                   unread=0, last="")
        db.session.add(t)
        db.session.flush()
    m = Message(thread_id=t.id, sender="AGENT", text=text,
                t=datetime.utcnow().strftime("%H:%M"))
    db.session.add(m)
    t.unread = (t.unread or 0) + 1
    t.last = text
    db.session.commit()
    deep_link = f"/provider/respond?pnr={pnr}&agency={DEMO_TERMINAL['agency'].replace(' ', '+')}"
    return jsonify({
        "status": "OK",
        "thread_id": t.id,
        "deep_link": deep_link,
        "message": "Support chat updated — the helpdesk sees your PNR context.",
    })


# --- Routes: Agency Portal (client updates Batch 3) ------------------------
#
# The third tier: agency admins (max 3 sub-users) get performance KPIs,
# issuance reporting with Excel export, visa lookup, sub-user management
# with a permission matrix, and the airline-policy/IATA reference page.

MAX_AGENCY_SUBUSERS = 3
PORTAL_PERM_KEYS = [
    ("reports", "View financial / issuance reports"),
    ("visa", "Use the visa requirement tool"),
    ("chat", "Use live chat with AERO-GUARD"),
    ("escalate", "Raise case escalations"),
]


def current_agency() -> Agency:
    """The logged-in portal user's agency, or 403 if the link is broken."""
    a = Agency.query.get(current_user.agency_id) if current_user.agency_id else None
    if a is None or a.deleted_at is not None:
        abort(403)
    return a


def write_portal_audit(action: str, target_type: str, target_id: str, *, note: str = "") -> None:
    """Audit rows from portal actions land under the agency's provider so
    the helpdesk sees them in the tenant-scoped audit log."""
    a = current_agency()
    db.session.add(AuditLog(
        provider_id=a.provider_id,
        actor_user_id=current_user.id,
        action=action,
        target_type=target_type,
        target_id=str(target_id),
        note=note or None,
    ))


def render_portal(template_name, active_nav, **context):
    a = current_agency()
    return render_template(
        template_name,
        agency=a,
        agency_tier=policy_tier(a.policy_level),
        active_nav=active_nav,
        **context,
    )


def _traffic_light(pct: float) -> str:
    """The spec's traffic-light bands: ≥90 green, 75–89 amber, <75 red."""
    if pct >= 90:
        return "green"
    if pct >= 75:
        return "amber"
    return "red"


def _chart_area_path(series, key, w=640, h=150, maxv=None):
    """SVG area path for the 30-day compliance chart (no JS chart lib)."""
    if not series:
        return ""
    maxv = maxv or 1
    n = len(series)
    pts = []
    for i, s in enumerate(series):
        x = round(i * (w / max(1, n - 1)), 1)
        y = round(h - (min(s[key], maxv) / maxv) * (h - 14), 1)
        pts.append(f"{x},{y}")
    return f"M0,{h} L" + " L".join(pts) + f" L{w},{h} Z"


@app.route("/portal")
def portal_dashboard():
    a = current_agency()
    tickets = TicketIssue.query.filter_by(agency_id=a.id).all()
    total = len(tickets)
    overridden = [t for t in tickets if t.overridden]
    compliance_pct = round(100.0 * (1 - len(overridden) / total), 1) if total else 100.0
    adm_avoided = round(sum(t.saved_amount or 0 for t in tickets))
    adm_incurred = round(sum(t.adm_amount or 0 for t in tickets))
    open_esc = (Escalation.query
                .filter_by(agency=a.name)
                .filter(Escalation.status != "RESOLVED").count())

    # 30-day series: compliant vs overridden issuance per day.
    today = datetime.utcnow().date()
    days = [today - timedelta(days=i) for i in range(29, -1, -1)]
    by_day = {d: {"ok": 0, "ovr": 0} for d in days}
    for t in tickets:
        d = (t.issued_at or datetime.utcnow()).date()
        if d in by_day:
            by_day[d]["ovr" if t.overridden else "ok"] += 1
    series = [{"label": d.strftime("%d %b"), "ok": by_day[d]["ok"], "ovr": by_day[d]["ovr"],
               "total": by_day[d]["ok"] + by_day[d]["ovr"]} for d in days]
    maxv = max(1, max(s["total"] for s in series))

    # Broadcast engine: same rows the provider posts once — zero-lag mirror.
    announcements = [
        {"source": b.source, "time": humanize(b.created_at), "tag": b.tag,
         "title": b.title, "kind": b.kind}
        for b in (Broadcast.query.filter_by(provider_id=a.provider_id)
                  .order_by(Broadcast.created_at.desc()).limit(6).all())
    ]
    return render_portal(
        "portal/dashboard.html", "DASHBOARD",
        total_tickets=total,
        compliance_pct=compliance_pct,
        compliance_light=_traffic_light(compliance_pct),
        adm_avoided=adm_avoided,
        adm_incurred=adm_incurred,
        open_escalations=open_esc,
        avg_ticketing="4.2m",
        series=series,
        chart_ok_path=_chart_area_path(series, "total", maxv=maxv),
        chart_ovr_path=_chart_area_path(series, "ovr", maxv=maxv),
        announcements=announcements,
    )


@app.route("/portal/reports")
def portal_reports():
    if not current_user.portal_can("reports"):
        return render_template("auth/403.html"), 403
    a = current_agency()
    filter_airline = request.args.get("airline", "ALL")
    q = (TicketIssue.query.filter_by(agency_id=a.id)
         .order_by(TicketIssue.issued_at.desc()))
    tickets = q.all()
    airlines = sorted({t.airline for t in tickets})
    rows = [t for t in tickets if filter_airline == "ALL" or t.airline == filter_airline]

    total_value = round(sum(t.amount or 0 for t in rows), 2)
    adm_exposure = round(sum(t.adm_amount or 0 for t in rows), 2)
    ignored = [t for t in rows if t.overridden]
    roi_saved = round(sum(t.saved_amount or 0 for t in rows), 2)
    return render_portal(
        "portal/reports.html", "REPORTS",
        rows=rows, airlines=airlines, filter_airline=filter_airline,
        total_value=total_value, adm_exposure=adm_exposure,
        ignored_count=len(ignored), roi_saved=roi_saved,
    )


@app.route("/portal/reports.xlsx")
def portal_reports_xlsx():
    if not current_user.portal_can("reports"):
        return render_template("auth/403.html"), 403
    a = current_agency()
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Issuance"
    head = Font(bold=True)
    headers = ["Ticket #", "PNR", "Passenger", "Airline", "Route",
               "Amount", "Currency", "Issued (UTC)", "Agent",
               "Overridden warning", "Override reason", "ADM incurred", "ADM avoided"]
    ws.append(headers)
    for c in ws[1]:
        c.font = head
    tickets = (TicketIssue.query.filter_by(agency_id=a.id)
               .order_by(TicketIssue.issued_at.desc()).all())
    for t in tickets:
        ws.append([
            t.ticket_no, t.pnr, t.pax_name, t.airline, t.route,
            round(t.amount or 0, 2), t.currency,
            t.issued_at.strftime("%Y-%m-%d %H:%M") if t.issued_at else "",
            t.agent, "YES" if t.overridden else "",
            t.override_reason or "", round(t.adm_amount or 0, 2),
            round(t.saved_amount or 0, 2),
        ])
    ws.append([])
    totals = ["TOTALS", "", "", "", "", round(sum(t.amount or 0 for t in tickets), 2),
              "", "", "", sum(1 for t in tickets if t.overridden), "",
              round(sum(t.adm_amount or 0 for t in tickets), 2),
              round(sum(t.saved_amount or 0 for t in tickets), 2)]
    ws.append(totals)
    for c in ws[ws.max_row]:
        c.font = head
    widths = [14, 10, 24, 8, 16, 10, 9, 17, 26, 18, 26, 13, 12]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = wdt

    buf = io.BytesIO()
    wb.save(buf)
    # Spec filename format: NabiTravel_Issuance_Report_2026-08-08.xlsx
    stamp = datetime.utcnow().strftime("%Y-%m-%d")
    fname = f"{a.name.replace(' ', '')}_Issuance_Report_{stamp}.xlsx"
    write_portal_audit("PORTAL_REPORT_EXPORT", "agency", a.id, note=fname)
    db.session.commit()
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---- Portal: My Profile (sub-user management, max 3) ----------------------

def _own_subuser_or_404(uid: str) -> User:
    u = User.query.get(uid)
    if u is None or u.agency_id != current_user.agency_id or u.role != "AGENCY_USER":
        abort(404)
    return u


def _perms_from_form(form) -> str:
    return json.dumps({key: bool(form.get(f"perm_{key}")) for key, _ in PORTAL_PERM_KEYS})


@app.route("/portal/profile")
def portal_profile():
    a = current_agency()
    subs = (User.query.filter_by(agency_id=a.id, role="AGENCY_USER")
            .order_by(User.created_at).all())
    sub_rows = []
    for s in subs:
        try:
            perms = json.loads(s.portal_perms) if s.portal_perms else {}
        except (ValueError, TypeError):
            perms = {}
        sub_rows.append({"u": s, "perms": perms})
    return render_portal(
        "portal/profile.html", "PROFILE",
        subs=sub_rows, max_subusers=MAX_AGENCY_SUBUSERS,
        perm_keys=PORTAL_PERM_KEYS,
        seats_left=MAX_AGENCY_SUBUSERS - len(subs),
    )


@app.route("/portal/profile/subusers/add", methods=["POST"])
@require_role("AGENCY_ADMIN")
def portal_subuser_add():
    a = current_agency()
    count = User.query.filter_by(agency_id=a.id, role="AGENCY_USER").count()
    if count >= MAX_AGENCY_SUBUSERS:
        flash(f"Sub-user limit reached — your subscription allows {MAX_AGENCY_SUBUSERS}.", "error")
        return redirect(url_for("portal_profile"))
    name = (request.form.get("name") or "").strip()
    email = (request.form.get("email") or "").strip().lower()
    if not name or not email or "@" not in email:
        flash("A name and a valid email are required.", "error")
        return redirect(url_for("portal_profile"))
    if User.query.filter(db.func.lower(User.email) == email).first():
        flash("That email already has an account.", "error")
        return redirect(url_for("portal_profile"))
    temp_password = secrets.token_urlsafe(8)
    u = User(
        id=f"AU-{random.randint(100, 999)}",
        provider_id=None,
        agency_id=a.id,
        name=name,
        email=email,
        role="AGENCY_USER",
        active=True,
        mfa=False,
        last_login="never",
        portal_perms=_perms_from_form(request.form),
    )
    u.set_password(temp_password)
    db.session.add(u)
    write_portal_audit("PORTAL_SUBUSER_ADD", "user", u.id, note=email)
    db.session.commit()
    sent = send_email(
        to=email,
        subject=f"Your AERO-GUARD portal access — {a.name}",
        text=(f"Hi {name},\n\n{current_user.name} added you to {a.name}'s AERO-GUARD portal.\n\n"
              f"Sign in: {url_for('login', _external=True)}\nEmail: {email}\n"
              f"Temporary password: {temp_password}\n\nPlease change it after first sign-in."),
    )
    if sent:
        flash(f"{name} added — credentials emailed to {email}.", "success")
    else:
        flash(f"{name} added. Temporary password (shown once): {temp_password}", "success")
    return redirect(url_for("portal_profile"))


@app.route("/portal/profile/subusers/<uid>/toggle", methods=["POST"])
@require_role("AGENCY_ADMIN")
def portal_subuser_toggle(uid):
    u = _own_subuser_or_404(uid)
    u.active = not u.active
    write_portal_audit("PORTAL_SUBUSER_TOGGLE", "user", u.id,
                       note=f"{u.email} → {'active' if u.active else 'deactivated'}")
    db.session.commit()
    flash(f"{u.name} {'re-activated' if u.active else 'deactivated — access revoked immediately'}.",
          "success" if u.active else "info")
    return redirect(url_for("portal_profile"))


@app.route("/portal/profile/subusers/<uid>/remove", methods=["POST"])
@require_role("AGENCY_ADMIN")
def portal_subuser_remove(uid):
    u = _own_subuser_or_404(uid)
    write_portal_audit("PORTAL_SUBUSER_REMOVE", "user", u.id, note=u.email)
    db.session.delete(u)
    db.session.commit()
    flash(f"{u.name} removed.", "success")
    return redirect(url_for("portal_profile"))


@app.route("/portal/profile/subusers/<uid>/perms", methods=["POST"])
@require_role("AGENCY_ADMIN")
def portal_subuser_perms(uid):
    u = _own_subuser_or_404(uid)
    u.portal_perms = _perms_from_form(request.form)
    write_portal_audit("PORTAL_SUBUSER_PERMS", "user", u.id, note=u.portal_perms)
    db.session.commit()
    flash(f"Permissions updated for {u.name}.", "success")
    return redirect(url_for("portal_profile"))


@app.route("/portal/profile/subusers/<uid>/reset", methods=["POST"])
@require_role("AGENCY_ADMIN")
def portal_subuser_reset(uid):
    u = _own_subuser_or_404(uid)
    temp_password = secrets.token_urlsafe(8)
    u.set_password(temp_password)
    write_portal_audit("PORTAL_SUBUSER_RESET", "user", u.id, note=u.email)
    db.session.commit()
    sent = send_email(
        to=u.email,
        subject="Your AERO-GUARD portal password was reset",
        text=f"Hi {u.name},\n\nYour temporary password: {temp_password}\nPlease change it after signing in.",
    )
    if sent:
        flash(f"New temporary password emailed to {u.email}.", "success")
    else:
        flash(f"New temporary password for {u.name} (shown once): {temp_password}", "success")
    return redirect(url_for("portal_profile"))


@app.route("/portal/profile/branding", methods=["POST"])
@require_role("AGENCY_ADMIN")
def portal_branding():
    """SB-10: choose how the AERO-GUARD trust badge renders on itineraries."""
    a = current_agency()
    style = request.form.get("badge_style")
    if style not in {"PROMINENT", "SUBTLE"}:
        style = "PROMINENT"
    a.badge_style = style
    write_portal_audit("PORTAL_BRANDING_SET", "agency", a.id, note=style)
    db.session.commit()
    flash(f"Itinerary trust badge set to {style.lower()} — applies to every #AG-generated document.", "success")
    return redirect(url_for("portal_profile"))


# ---- Portal: airline policies & IATA guidelines (kept current) -------------

AIRLINE_POLICY_FEED = [
    {"airline": "EK", "area": "Name changes", "effective": "Current",
     "policy": "No name changes after PNR creation — cancel and rebook. No name changes on issued tickets."},
    {"airline": "LH", "area": "Name changes", "effective": "Current",
     "policy": "Corrections up to 3 characters permitted before ticketing with documentary proof."},
    {"airline": "ET", "area": "Baggage", "effective": "01 Aug 2026",
     "policy": "Economy piece concept 2PC · 23kg on African routes; excess charged per piece."},
    {"airline": "SA", "area": "Point of commencement", "effective": "Current",
     "policy": "Tickets must be issued in the country of commencement — violations attract USD 300 ADM."},
    {"airline": "QR", "area": "ADM disputes", "effective": "Current",
     "policy": "Evidence pack required within 14 days of ADM issue; disputes via CASS portal only."},
    {"airline": "EK", "area": "Infants", "effective": "Current",
     "policy": "INF must be associated to an adult on the same PNR; maximum 1 INF per ADT."},
]

IATA_GUIDELINES = [
    {"ref": "Reso 830a", "topic": "Ticketing time limits",
     "note": "Agents must observe carrier TTLs; auto-cancellation applies on expiry."},
    {"ref": "Reso 890", "topic": "Card sales rules",
     "note": "Merchant of record rules for BSP card sales; CVV verification mandatory."},
    {"ref": "Reso 852", "topic": "Designation / selection of validating carrier",
     "note": "The validating carrier must participate in all segments' interline agreements."},
    {"ref": "TIMATIC", "topic": "Travel document verification",
     "note": "Verify passport validity (6 months), visas and transit requirements before issuance."},
]


@app.route("/portal/policies")
def portal_policies():
    docs = PolicyDoc.query.all()
    return render_portal(
        "portal/policies.html", "POLICIES",
        airline_policies=AIRLINE_POLICY_FEED,
        iata_guidelines=IATA_GUIDELINES,
        docs=[{"cat": d.cat, "name": d.name, "v": d.v} for d in docs],
    )


# ---- Provider: broadcast engine (single source → everywhere) ---------------

@app.route("/provider/broadcasts/new", methods=["POST"])
@require_role("ADMIN")
def provider_broadcast_new():
    title = (request.form.get("title") or "").strip()
    if not title:
        flash("A broadcast needs a title.", "error")
        return redirect(url_for("provider_dashboard"))
    kind = request.form.get("kind") or "INDUSTRY"
    tag = {"MAINTENANCE": "maint", "PROMO": "promo"}.get(kind, request.form.get("tag") or "policy")
    b = Broadcast(
        provider_id=current_provider_id(),
        kind=kind,
        source=(request.form.get("source") or "AERO-GUARD").strip()[:40],
        tag=tag,
        title=title[:255],
        author=current_user.name,
    )
    db.session.add(b)
    write_audit("BROADCAST_POST", "broadcast", title[:40], note=kind)
    # Email push to every agency admin of this tenant (simulated when SMTP
    # is unconfigured — the mailer logs it).
    notified = 0
    for a in all_agencies():
        if a.admin_email:
            send_email(
                to=a.admin_email,
                subject=f"[AERO-GUARD {kind.title()}] {title[:80]}",
                text=(f"Hello {a.name},\n\n{title}\n\n"
                      "This notice is also on your Agency Portal dashboard.\n\n— AERO-GUARD"),
            )
            notified += 1
    db.session.commit()
    flash(f"Broadcast published — live on the provider feed and {notified} agency portal(s); "
          f"email push {'sent' if mail_is_configured() else 'logged (SMTP not configured)'}.",
          "success")
    return redirect(url_for("provider_dashboard"))


@app.route("/provider/broadcasts/<int:bid>/delete", methods=["POST"])
@require_role("ADMIN")
def provider_broadcast_delete(bid):
    b = Broadcast.query.get(bid)
    if b is None or b.provider_id != current_provider_id():
        abort(404)
    db.session.delete(b)
    write_audit("BROADCAST_DELETE", "broadcast", str(bid), note=b.title[:60])
    db.session.commit()
    flash("Broadcast removed from all feeds.", "success")
    return redirect(url_for("provider_dashboard"))


# ---- Portal: live chat with greeting flow (strictly separated from
#      case escalation, per the client's spec) ------------------------------

def _portal_thread(a: Agency, create: bool = True) -> Thread | None:
    """The provider-side support thread this agency's chat binds to."""
    t = (Thread.query
         .filter_by(provider_id=a.provider_id, agency=a.name)
         .first())
    if t is None and create:
        t = Thread(id=f"T-{random.randint(100, 999)}",
                   provider_id=a.provider_id, agency=a.name,
                   agent=current_user.name, unread=0, last="")
        db.session.add(t)
        db.session.flush()
    return t


@app.route("/portal/chat")
def portal_chat():
    if not current_user.portal_can("chat"):
        return render_template("auth/403.html"), 403
    a = current_agency()
    greeted = session.get(f"portal_chat_greeted_{a.id}")
    messages = []
    if greeted:
        t = _portal_thread(a, create=False)
        if t:
            messages = [{"id": m.id, "from": m.sender, "text": m.text, "t": m.t}
                        for m in t.messages]
    return render_portal(
        "portal/chat.html", "CHAT",
        greeted=bool(greeted), messages=messages,
        greet_defaults={
            "agent": current_user.name, "agency": a.name,
            "pcc": a.pcc, "country": a.country,
        },
    )


@app.route("/portal/chat/greet", methods=["POST"])
def portal_chat_greet():
    if not current_user.portal_can("chat"):
        return render_template("auth/403.html"), 403
    a = current_agency()
    agent = (request.form.get("agent") or current_user.name).strip()
    pcc = (request.form.get("pcc") or a.pcc).strip().upper()
    country = (request.form.get("country") or a.country).strip().upper()
    session[f"portal_chat_greeted_{a.id}"] = {"agent": agent, "pcc": pcc, "country": country}
    t = _portal_thread(a)
    m = Message(thread_id=t.id, sender="AGENT",
                text=f"Chat opened · {agent} · {a.name} · PCC {pcc} · {country}",
                t=datetime.utcnow().strftime("%H:%M"))
    db.session.add(m)
    t.unread = (t.unread or 0) + 1
    t.last = m.text
    db.session.commit()
    return redirect(url_for("portal_chat"))


@app.route("/portal/chat/send", methods=["POST"])
def portal_chat_send():
    if not current_user.portal_can("chat"):
        return render_template("auth/403.html"), 403
    a = current_agency()
    text = (request.form.get("text") or "").strip()
    if text:
        t = _portal_thread(a)
        m = Message(thread_id=t.id, sender="AGENT", text=text[:500],
                    t=datetime.utcnow().strftime("%H:%M"))
        db.session.add(m)
        t.unread = (t.unread or 0) + 1
        t.last = text[:255]
        db.session.commit()
    return redirect(url_for("portal_chat"))


@app.route("/portal/chat.json")
def portal_chat_json():
    """Light polling endpoint so helpdesk replies appear without a reload."""
    if not current_user.portal_can("chat"):
        return jsonify({"messages": []}), 403
    a = current_agency()
    after = request.args.get("after", type=int, default=0)
    t = _portal_thread(a, create=False)
    msgs = []
    if t:
        msgs = [{"id": m.id, "from": m.sender, "text": m.text, "t": m.t}
                for m in t.messages if m.id > after]
    return jsonify({"messages": msgs})


# ---- Portal: case escalation module (formal, typed, tier-routed) ----------

@app.route("/portal/escalate", methods=["GET", "POST"])
def portal_escalate():
    if not current_user.portal_can("escalate"):
        return render_template("auth/403.html"), 403
    a = current_agency()

    if request.method == "POST":
        subject = (request.form.get("subject") or "").strip()
        if not subject:
            flash("A subject is required.", "error")
            return redirect(url_for("portal_escalate"))
        category = request.form.get("category") or "GENERAL"
        if category not in ESCALATION_TIERS:
            category = "GENERAL"
        priority = request.form.get("priority") or "MED"
        new_id = f"ESC-{random.randint(9000, 9999)}"
        e = Escalation(
            id=new_id,
            provider_id=a.provider_id,
            agency=a.name,
            pnr=(request.form.get("pnr") or "—").upper()[:20],
            subject=subject[:255],
            level="L1",
            category=category,
            priority=priority if priority in {"HIGH", "MED", "LOW"} else "MED",
            opened="just now",
            status="OPEN",
            sla="24 hr left",
        )
        e.sla_due_at = datetime.utcnow() + timedelta(hours=4 if priority == "HIGH" else 24)
        db.session.add(e)
        write_portal_audit("PORTAL_ESCALATION_CREATE", "escalation", new_id,
                           note=f"{category} · {priority} · {subject[:60]}")
        db.session.commit()
        flash(f"Case {new_id} filed — routed to {ESCALATION_TIERS[category]}.", "success")
        return redirect(url_for("portal_escalate"))

    own = (Escalation.query
           .filter_by(provider_id=a.provider_id, agency=a.name)
           .order_by(Escalation.created_at.desc()).all())
    return render_portal(
        "portal/escalate.html", "ESCALATE",
        cases=[escalation_to_dict(e) for e in own],
        tiers=ESCALATION_TIERS,
    )


# --- Routes: provider dashboard -------------------------------------------

@app.route("/provider")
def provider_dashboard():
    # Tenant-scoped counters from real data.
    own_agencies = all_agencies()
    adms_prevented = sum(a.month_adms for a in own_agencies)
    alerts = Alert.query.all()  # catalog data, shared across tenants
    # Pending issues are catalog/seeded — filter to the agencies we own.
    own_agency_names = {a.name for a in own_agencies}
    pending = [p for p in PendingIssue.query.all() if p.agency in own_agency_names]
    # Demo: surface a system-wide ADM spike when several agencies hit the same
    # rule/route in a short window — the kind of pattern the helpdesk broadcasts on.
    adm_spike = {
        "agencies": 4,
        "rule": "POINT OF COMMENCEMENT",
        "route": "HRE-JNB",
        "airline": "SA",
        "window": "24h",
    }
    # Broadcast engine (Batch 4): the feed is now DB-backed — posted once
    # here, mirrored instantly on every Agency Portal + email push.
    industry_feed = [
        {"id": b.id, "source": b.source, "time": humanize(b.created_at),
         "tag": b.tag, "title": b.title, "kind": b.kind}
        for b in (tenant_q(Broadcast).order_by(Broadcast.created_at.desc()).limit(8).all())
    ]

    # Predictive bypass alerts (Batch 5, PH-8): flag any user who repeatedly
    # issued against warnings so the helpdesk can intervene before an ADM.
    bypass_counts: dict[str, int] = {}
    own_ids = {a.id for a in own_agencies}
    for t in TicketIssue.query.filter_by(provider_id=current_provider_id(), overridden=True).all():
        if t.agency_id in own_ids:
            bypass_counts[t.agent] = bypass_counts.get(t.agent, 0) + 1
    bypass_alerts = sorted(
        ({"agent": agent, "count": n} for agent, n in bypass_counts.items() if n >= 3),
        key=lambda r: -r["count"],
    )
    return render_provider(
        "provider/dashboard.html",
        "DASHBOARD",
        adm_spike=adm_spike,
        bypass_alerts=bypass_alerts,
        industry_feed=industry_feed,
        adms_prevented=adms_prevented,
        dollar_saved=adms_prevented * 350,
        alerts=[{"id": a.id, "severity": a.severity, "source": a.source, "title": a.title,
                 "time": a.time, "ongoing": a.ongoing, "impacted_agencies": a.impacted_agencies} for a in alerts],
        pending_issues=[{"id": p.id, "agency": p.agency, "type": p.type, "summary": p.summary,
                         "age": p.age, "priority": p.priority} for p in pending],
    )


# --- Routes: agencies -----------------------------------------------------

SORT_KEYS = {
    "name": lambda a: a["name"].lower(),
    "seats": lambda a: (a["used_seats"] / a["seats"]) if a["seats"] else 0,
    "adms": lambda a: a["month_adms"],
    "lastActive": lambda a: a["last_active"],
}


def _render_agencies_page(prefill: dict | None = None):
    """Render the agencies list. ``prefill`` re-opens the provisioning modal
    with previously entered values (used by the Review screen's [Edit])."""
    filter_gds = request.args.get("gds", "ALL")
    filter_country = request.args.get("country", "ALL")
    filter_policy = request.args.get("policy", "ALL")
    sort_by = request.args.get("sort", "name")
    sort_dir = request.args.get("dir", "asc")

    rows = [agency_to_dict(a) for a in all_agencies()]
    rows = [
        a for a in rows
        if (filter_gds == "ALL" or a["gds"] == filter_gds)
        and (filter_country == "ALL" or a["country"] == filter_country)
        and (filter_policy == "ALL" or a["policy_tier"] == filter_policy)
    ]
    rows.sort(key=SORT_KEYS.get(sort_by, SORT_KEYS["name"]), reverse=(sort_dir == "desc"))

    def sort_url(key):
        next_dir = "desc" if (sort_by == key and sort_dir == "asc") else "asc"
        return url_for("provider_agencies", gds=filter_gds, country=filter_country,
                       policy=filter_policy, sort=key, dir=next_dir)

    return render_provider(
        "provider/agencies.html", "AGENCIES",
        agencies=rows, countries=COUNTRIES, policy_templates=POLICY_TEMPLATES,
        regions=REGIONS, prefill=prefill,
        filter_gds=filter_gds, filter_country=filter_country, filter_policy=filter_policy,
        sort_by=sort_by, sort_dir=sort_dir, sort_url=sort_url,
        querystring=request.query_string.decode(),
    )


@app.route("/provider/agencies")
def provider_agencies():
    return _render_agencies_page()


# --- Provisioning workflow (client updates Batch 1) -----------------------
#
# Data entry (modal) → Review & Confirm screen → agency saved ACTIVE →
# Email Notification Hub with pre-generated welcome emails for the
# manager + up to 3 consultants.

MAX_CONSULTANT_SEATS = 3


def _provision_form(form) -> dict:
    """Normalize the provisioning form fields into one dict that the
    review / edit / confirm steps all share."""
    members = []
    mgr_name = (form.get("manager_name") or "").strip()
    mgr_email = (form.get("manager_email") or "").strip().lower()
    if mgr_name or mgr_email:
        members.append({"name": mgr_name, "email": mgr_email, "role": "MANAGER"})
    for i in range(1, MAX_CONSULTANT_SEATS + 1):
        n = (form.get(f"consultant{i}_name") or "").strip()
        e = (form.get(f"consultant{i}_email") or "").strip().lower()
        if n or e:
            members.append({"name": n, "email": e, "role": "CONSULTANT"})
    try:
        seats = int(form.get("seats") or 10)
    except ValueError:
        seats = 0
    return {
        "name": (form.get("name") or "").strip(),
        "pcc": (form.get("pcc") or "").strip().upper(),
        "gds": form.get("gds") or "1G",
        "country": form.get("country") or "ZW",
        "region": form.get("region") or REGIONS[0],
        "seats": seats,
        "policy": form.get("policy") or "Gold",
        "mode": form.get("mode") or "FULL",
        "members": members,
    }


def _provision_errors(data: dict) -> list[str]:
    errs = []
    if not data["name"]:
        errs.append("Agency name is required.")
    if not data["pcc"]:
        errs.append("PCC is required.")
    if data["seats"] < 1:
        errs.append("Seat count must be at least 1.")
    managers = [m for m in data["members"] if m["role"] == "MANAGER"]
    consultants = [m for m in data["members"] if m["role"] == "CONSULTANT"]
    if not managers:
        errs.append("The agency manager's name and email are required.")
    if len(consultants) > MAX_CONSULTANT_SEATS:
        errs.append(f"At most {MAX_CONSULTANT_SEATS} consultant sub-users are allowed.")
    for m in data["members"]:
        if not m["name"] or not m["email"] or "@" not in m["email"]:
            errs.append(f"Each {m['role'].lower()} entry needs both a name and a valid email.")
            break
    return errs


@app.route("/provider/agencies/provision/review", methods=["POST"])
@require("agency:provision")
def provision_review():
    """Step 2 of the workflow: show everything just entered for sign-off."""
    data = _provision_form(request.form)
    errors = _provision_errors(data)
    if errors:
        for e in errors:
            flash(e, "error")
        return _render_agencies_page(prefill=data)
    return render_provider("provider/provision_review.html", "AGENCIES", data=data)


@app.route("/provider/agencies/provision/edit", methods=["POST"])
@require("agency:provision")
def provision_edit():
    """[Edit] on the review screen — reopen the modal with values intact."""
    return _render_agencies_page(prefill=_provision_form(request.form))


@app.route("/provider/agencies/provision", methods=["POST"])
@require("agency:provision")
def provision_agency():
    """[Confirm Provisioning] — persist the agency ACTIVE + whitelist, then
    hand off to the Email Notification Hub."""
    data = _provision_form(request.form)
    errors = _provision_errors(data)
    if errors:
        for e in errors:
            flash(e, "error")
        return _render_agencies_page(prefill=data)

    new_id = f"AG-{random.randint(1000, 9999)}"
    manager = next(m for m in data["members"] if m["role"] == "MANAGER")
    status = "TRIAL" if data["mode"] == "TRIAL" else "ACTIVE"
    db.session.add(Agency(
        id=new_id,
        provider_id=current_provider_id(),
        name=data["name"],
        pcc=data["pcc"] or "XXXX",
        gds=data["gds"],
        country=data["country"],
        region=data["region"],
        seats=data["seats"],
        used_seats=0,
        status=status,
        month_adms=0,
        last_active="—",
        policy_level=data["policy"],
        admin_email=manager["email"],
    ))
    db.session.flush()  # agency row must exist before members reference it

    temp_creds: dict[str, str] = {}
    for m in data["members"]:
        row = AgencyMember(agency_id=new_id, name=m["name"], email=m["email"], member_role=m["role"])
        db.session.add(row)
        db.session.flush()
        # One-time temporary password: shown on the hub for this session
        # only, never persisted. Real portal logins arrive with the
        # Agency Portal tier.
        temp_creds[str(row.id)] = secrets.token_urlsafe(8)

    write_audit("AGENCY_PROVISION", "agency", new_id,
                note=f"{data['name']} · {status} · {len(data['members'])} member(s)")
    db.session.commit()
    session[f"welcome_creds_{new_id}"] = temp_creds
    flash(f"{data['name']} provisioned and marked {status}.", "success")
    return redirect(url_for("agency_welcome_hub", agency_id=new_id))


def _welcome_email_for(member: AgencyMember, agency: Agency, temp_password: str | None):
    """Build the (subject, body) pair for one member's welcome email."""
    login_url = url_for("login", _external=True)
    tier = policy_tier(agency.policy_level).title()
    lines = [
        f"Hi {member.name},",
        "",
        f"{agency.name} has been provisioned on AERO-GUARD "
        f"({tier} tier · {agency.seats} seats · GDS {agency.gds}).",
        "",
        f"Sign in:            {login_url}",
        f"Your login email:   {member.email}",
        f"Temporary password: {temp_password or '(issued separately by your administrator)'}",
        "",
        "First steps:",
        "  1. Sign in and change your password.",
        "  2. Enrol multi-factor authentication from your account menu (recommended).",
    ]
    if member.member_role == "CONSULTANT":
        lines.append("  3. Open the consultant terminal and press the AERO-GUARD Smart Button, or type #AG.")
    lines += [
        "",
        "Need help? Use the live chat in your portal or reply to this email.",
        "",
        "— AERO-GUARD Provisioning",
    ]
    subject = f"Welcome to AERO-GUARD — {agency.name} is now active"
    return subject, "\n".join(lines)


@app.route("/provider/agencies/<agency_id>/welcome-hub")
@require("agency:provision")
def agency_welcome_hub(agency_id):
    """Step 4: the Email Notification Hub for one agency's whitelist."""
    a = get_owned_or_404(Agency, agency_id)
    creds = session.get(f"welcome_creds_{agency_id}", {})
    members = (AgencyMember.query.filter_by(agency_id=agency_id)
               .order_by(AgencyMember.member_role.desc(), AgencyMember.id).all())
    cards = []
    for m in members:
        subject, body = _welcome_email_for(m, a, creds.get(str(m.id)))
        cards.append({
            "id": m.id, "name": m.name, "email": m.email, "role": m.member_role,
            "subject": subject, "body": body,
            "sent": humanize(m.welcome_sent_at) if m.welcome_sent_at else None,
        })
    return render_provider(
        "provider/welcome_hub.html", "AGENCIES",
        agency=agency_to_dict(a), cards=cards,
        mail_configured=mail_is_configured(), has_creds=bool(creds),
    )


@app.route("/provider/agencies/<agency_id>/welcome-hub/send", methods=["POST"])
@require("agency:provision")
def agency_welcome_send(agency_id):
    a = get_owned_or_404(Agency, agency_id)
    creds = session.get(f"welcome_creds_{agency_id}", {})
    member_id = request.form.get("member_id") or "ALL"
    q = AgencyMember.query.filter_by(agency_id=agency_id)
    if member_id == "ALL":
        members = q.all()
    else:
        members = [q.filter_by(id=int(member_id)).first_or_404()]

    delivered = 0
    for m in members:
        subject, body = _welcome_email_for(m, a, creds.get(str(m.id)))
        ok = send_email(to=m.email, subject=subject, text=body)
        m.welcome_sent_at = datetime.utcnow()
        delivered += 1 if ok else 0
        write_audit("AGENCY_WELCOME_SENT", "agency", a.id,
                    note=f"{m.email} ({'delivered' if ok else 'logged — SMTP unconfigured'})")
    db.session.commit()

    if mail_is_configured():
        flash(f"Welcome email sent to {delivered} recipient(s).", "success")
    else:
        flash(f"SMTP not configured — {len(members)} welcome email(s) logged. "
              "The preview below is exactly what will be sent once email is set up.", "info")
    return redirect(url_for("agency_welcome_hub", agency_id=agency_id))


@app.route("/provider/agencies/<agency_id>/toggle-suspend", methods=["POST"])
@require("agency:suspend")
def toggle_suspend_agency(agency_id):
    a = get_owned_or_404(Agency, agency_id)
    a.status = "ACTIVE" if a.status == "SUSPENDED" else "SUSPENDED"
    write_audit("AGENCY_TOGGLE_SUSPEND", "agency", a.id, note=a.status)
    db.session.commit()
    return redirect(request.referrer or url_for("provider_agencies"))


@app.route("/provider/agencies/<agency_id>/delete", methods=["POST"])
@require("agency:delete")
def delete_agency(agency_id):
    a = get_owned_or_404(Agency, agency_id)
    a.deleted_at = datetime.utcnow()  # soft delete
    write_audit("AGENCY_DELETE", "agency", a.id, note=a.name)
    db.session.commit()
    return redirect(request.referrer or url_for("provider_agencies"))


@app.route("/provider/agencies/bulk", methods=["POST"])
def bulk_agency_action():
    action = request.form.get("action")
    ids = set(request.form.getlist("ids"))
    if not ids:
        return redirect(request.referrer or url_for("provider_agencies"))
    needed = "agency:bulk_delete" if action == "DELETE" else "agency:bulk_other"
    if not can(current_user, needed):
        return render_template("auth/403.html"), 403
    # Tenant-scope the bulk — silently ignore IDs that aren't ours.
    targets = (
        tenant_q(Agency)
        .filter(Agency.id.in_(ids))
        .all()
    )
    if action == "DELETE":
        now = datetime.utcnow()
        for a in targets:
            a.deleted_at = now
            write_audit("AGENCY_DELETE", "agency", a.id, note="bulk")
    elif action in ("SUSPEND", "REACTIVATE"):
        new_status = "SUSPENDED" if action == "SUSPEND" else "ACTIVE"
        for a in targets:
            a.status = new_status
            write_audit("AGENCY_TOGGLE_SUSPEND", "agency", a.id, note=f"bulk:{new_status}")
    db.session.commit()
    return redirect(request.referrer or url_for("provider_agencies"))


# --- Routes: helpdesk users ----------------------------------------------

@app.route("/provider/users")
def provider_users():
    users = (
        tenant_q(User)
        .filter(User.role != "CONSULTANT")
        .order_by(User.id)
        .all()
    )
    return render_provider(
        "provider/users.html", "USERS",
        users=[{"id": u.id, "name": u.name, "email": u.email, "role": u.role,
                "active": u.active, "mfa": u.mfa,
                "last_login": humanize(u.last_login_at) if u.last_login_at else "never"}
               for u in users],
        role_hints=ROLE_HINTS,
    )


@app.route("/provider/users/invite", methods=["POST"])
@require("user:invite")
def invite_user():
    import secrets as _secrets

    new_id = f"U-{random.randint(10, 99)}"
    email = (request.form.get("email") or f"{new_id.lower()}@example.com").lower()
    # Enforce global email uniqueness without leaking who already owns it.
    if User.query.filter(db.func.lower(User.email) == email).first():
        flash("That email is already registered.", "info")
        return redirect(url_for("provider_users"))
    name = request.form.get("name") or "Unnamed"
    role = request.form.get("role") or "L1"
    if role not in ("ADMIN", "L2", "L1"):
        role = "L1"

    # Invite = create the user with a first-password token (reuses the
    # reset flow) and email them the activation link. Token lives 7
    # days so the invite doesn't die overnight.
    token = _secrets.token_urlsafe(24)
    new_user = User(
        id=new_id,
        provider_id=current_provider_id(),
        name=name,
        email=email,
        role=role,
        active=True,
        mfa=False,
        last_login="never",
        reset_token=token,
        reset_expires=datetime.utcnow() + timedelta(days=7),
    )
    db.session.add(new_user)
    write_audit("USER_INVITE", "user", new_id, note=email)
    db.session.commit()

    link = url_for("reset_password", token=token, _external=True)
    inviter = getattr(current_user, "name", "your team")
    provider = Provider.query.get(current_provider_id())
    provider_name = provider.name if provider else "AERO-GUARD"
    email_sent = send_email(
        to=email,
        subject=f"You're invited to {provider_name} on AERO-GUARD",
        text=(
            f"Hi {name},\n\n"
            f"{inviter} has added you to {provider_name} on AERO-GUARD "
            f"as a {role} user.\n\n"
            "Set your password using the link below (valid 7 days):\n\n"
            f"{link}\n\n"
            "— AERO-GUARD"
        ),
        html=(
            f"<p>Hi {name},</p>"
            f"<p><strong>{inviter}</strong> has added you to <strong>{provider_name}</strong> "
            f"on AERO-GUARD as a <strong>{role}</strong> user.</p>"
            "<p>Set your password using the link below (valid 7 days):</p>"
            f'<p><a href="{link}">{link}</a></p>'
            "<p>&mdash; AERO-GUARD</p>"
        ),
    )
    if email_sent:
        flash(f"Invite sent to {email}.", "success")
    else:
        # Dev mode — surface the link so the admin can hand it over.
        flash(
            f"User created. Email isn't configured — share this activation "
            f"link with {email} (valid 7 days): {link}",
            "info",
        )
    return redirect(url_for("provider_users"))


@app.route("/provider/users/<user_id>/toggle-active", methods=["POST"])
@require("user:toggle_active")
def toggle_user_active(user_id):
    u = get_owned_or_404(User, user_id)
    if u.id == current_user.id:
        # Admins disabling themselves would lock them out — refuse.
        flash("You can't disable your own account.", "info")
        return redirect(request.referrer or url_for("provider_users"))
    u.active = not u.active
    write_audit("USER_TOGGLE_ACTIVE", "user", u.id, note="active" if u.active else "disabled")
    db.session.commit()
    return redirect(request.referrer or url_for("provider_users"))


@app.route("/provider/users/<user_id>/remove", methods=["POST"])
@require("user:remove")
def remove_user(user_id):
    u = get_owned_or_404(User, user_id)
    if u.id == current_user.id:
        flash("You can't remove your own account.", "info")
        return redirect(request.referrer or url_for("provider_users"))
    write_audit("USER_REMOVE", "user", u.id, note=u.email)
    db.session.delete(u)
    db.session.commit()
    return redirect(request.referrer or url_for("provider_users"))


@app.route("/provider/users/<user_id>/reset-link", methods=["POST"])
@require("user:reset_password")
def issue_reset_link(user_id):
    """Admin recovery: mint a fresh reset token for a user and either email
    it or surface it inline (when SMTP isn't configured), so an admin can
    help someone whose email is unreachable."""
    import secrets as _secrets

    u = get_owned_or_404(User, user_id)
    token = _secrets.token_urlsafe(24)
    u.reset_token = token
    u.reset_expires = datetime.utcnow() + timedelta(hours=2)
    db.session.commit()
    link = url_for("reset_password", token=token, _external=True)

    email_sent = send_email(
        to=u.email,
        subject="Your AERO-GUARD password was reset by an admin",
        text=(
            f"Hi {u.name},\n\n"
            f"An administrator ({getattr(current_user, 'name', 'AERO-GUARD')}) "
            "issued a password-reset link for your account. Follow the link "
            "below within the next 2 hours to set a new password:\n\n"
            f"{link}\n\n"
            "If you didn't expect this, contact your provider admin.\n\n"
            "— AERO-GUARD"
        ),
        html=(
            f"<p>Hi {u.name},</p>"
            f"<p>An administrator (<strong>{getattr(current_user, 'name', 'AERO-GUARD')}</strong>) "
            "issued a password-reset link for your account. Follow the link "
            "below within the next 2 hours to set a new password:</p>"
            f'<p><a href="{link}">{link}</a></p>'
            "<p>If you didn't expect this, contact your provider admin.</p>"
            "<p>&mdash; AERO-GUARD</p>"
        ),
    )
    write_audit(
        "USER_RESET_LINK_ISSUED", "user", u.id,
        note=f"emailed={email_sent}",
    )
    db.session.commit()
    if email_sent:
        flash(f"Reset link emailed to {u.email} (valid 2 hours).", "success")
    else:
        # Dev / SMTP-unset — hand the link to the admin so they can share it out-of-band.
        flash(
            f"Reset link for {u.email} (valid 2 hours) — email isn't configured, "
            f"share this URL out-of-band: {link}",
            "info",
        )
    return redirect(request.referrer or url_for("provider_users"))


@app.route("/provider/users/<user_id>/audit.json")
@require("audit:view")
def user_audit_json(user_id):
    """Return the audit trail for a single user — rows where they were the
    actor OR the target. Used by the "Audit" button modal on the Users page.
    Tenant-scoped through get_owned_or_404."""
    u = get_owned_or_404(User, user_id)
    try:
        limit = min(max(int(request.args.get("limit", 100)), 1), 500)
    except ValueError:
        limit = 100

    rows = (
        AuditLog.query
        .filter(AuditLog.provider_id == current_provider_id())
        .filter(db.or_(
            AuditLog.actor_user_id == u.id,
            db.and_(AuditLog.target_type == "user", AuditLog.target_id == u.id),
        ))
        .order_by(AuditLog.id.desc())
        .limit(limit)
        .all()
    )
    # Resolve actor names once — avoids N+1 lookups when rendering.
    actor_ids = {r.actor_user_id for r in rows if r.actor_user_id}
    actors = (
        {a.id: a.name for a in User.query.filter(User.id.in_(actor_ids)).all()}
        if actor_ids else {}
    )
    entries = [{
        "id": r.id,
        "time": r.created_at.strftime("%Y-%m-%d %H:%M UTC") if r.created_at else "",
        "action": r.action,
        "direction": "did" if r.actor_user_id == u.id else "on",
        "actor": actors.get(r.actor_user_id) or ("system" if not r.actor_user_id else r.actor_user_id),
        "target_type": r.target_type or "",
        "target_id": r.target_id or "",
        "note": r.note or "",
    } for r in rows]
    return jsonify({
        "user": {"id": u.id, "name": u.name, "email": u.email},
        "count": len(entries),
        "limit": limit,
        "entries": entries,
    })


@app.route("/provider/users/<user_id>/reset-mfa", methods=["POST"])
@require("user:reset_mfa")
def reset_user_mfa(user_id):
    """Admin recovery: clear a user's MFA enrollment so they can sign in
    with just their password and re-enrol from a device they control."""
    u = get_owned_or_404(User, user_id)
    if not u.mfa_secret and not u.mfa:
        flash(f"{u.name} doesn't have MFA enabled.", "info")
        return redirect(request.referrer or url_for("provider_users"))
    u.mfa_secret = None
    u.mfa_backup_codes = None
    u.mfa = False
    write_audit("USER_MFA_RESET", "user", u.id, note=u.email)
    db.session.commit()
    flash(
        f"MFA cleared for {u.email}. They can now sign in with password "
        "only and re-enrol from their device.",
        "success",
    )
    return redirect(request.referrer or url_for("provider_users"))


# --- Routes: audits ------------------------------------------------------

def _audit_rows(filter_agency: str, sort_health: bool) -> list[dict]:
    """Per-agency audit breakdown — shared by the page and the CSV export."""
    rows = []
    for a in all_agencies():
        if filter_agency != "ALL" and a.name != filter_agency:
            continue
        adms = a.month_adms
        health = "AT-RISK" if adms > 10 else ("WATCH" if adms > 3 else "HEALTHY")
        rows.append({
            "agency": a.name, "pcc": a.pcc, "adms": adms,
            "saved": adms * 1800 + 4000, "lost": adms * 400,
            "health": health, "trend": "↑" if adms > 5 else "↓",
            "tier": policy_tier(a.policy_level),
        })
    if sort_health:
        order = {"AT-RISK": 0, "WATCH": 1, "HEALTHY": 2}
        rows.sort(key=lambda r: order[r["health"]])
    return rows


@app.route("/provider/audits.csv")
def provider_audits_csv():
    range_ = request.args.get("range", "WEEK")
    filter_agency = request.args.get("agency", "ALL")
    sort_health = request.args.get("sortHealth") == "1"
    rows = _audit_rows(filter_agency, sort_health)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Agency", "PCC", "Policy tier", "ADMs", "Saved (USD)", "Lost (USD)", "Health", "Trend"])
    for r in rows:
        writer.writerow([r["agency"], r["pcc"], r["tier"], r["adms"], r["saved"], r["lost"], r["health"], r["trend"]])

    stamp = datetime.utcnow().strftime("%Y-%m-%d")
    scope = "all-agencies" if filter_agency == "ALL" else filter_agency.replace(" ", "-")
    filename = f"AERO-GUARD_ADM-Audit_{scope}_{range_}_{stamp}.csv"
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/provider/audits")
def provider_audits():
    range_ = request.args.get("range", "WEEK")
    filter_agency = request.args.get("agency", "ALL")
    sort_health = request.args.get("sortHealth") == "1"
    drill = request.args.get("drill")

    agencies = all_agencies()
    rows = _audit_rows(filter_agency, sort_health)

    # --- Risk intelligence (Batch 5, PH-7) — computed from real issuance ---
    tickets = TicketIssue.query.filter_by(provider_id=current_provider_id()).all()
    agency_names_by_id = {a.id: a.name for a in agencies}

    # ADM & loss mitigation by airline; SYSTEM-WIDE when overrides span
    # more than one agency (vs an isolated single-agency trend).
    by_airline: dict[str, dict] = {}
    for t in tickets:
        d = by_airline.setdefault(t.airline, {"airline": t.airline, "tickets": 0, "overrides": 0,
                                              "adm": 0.0, "saved": 0.0, "agencies": set()})
        d["tickets"] += 1
        d["saved"] += t.saved_amount or 0
        if t.overridden:
            d["overrides"] += 1
            d["adm"] += t.adm_amount or 0
            d["agencies"].add(t.agency_id)
    airline_risk = sorted(
        ({**d, "agencies": len(d["agencies"]),
          "scope": "SYSTEM-WIDE" if len(d["agencies"]) > 1 else ("SINGLE-AGENCY" if d["agencies"] else "—")}
         for d in by_airline.values()),
        key=lambda r: (-r["adm"], -r["overrides"]),
    )

    # Top vs struggling agencies by compliance rate.
    by_agency: dict[str, dict] = {}
    for t in tickets:
        d = by_agency.setdefault(t.agency_id, {"agency": agency_names_by_id.get(t.agency_id, t.agency_id),
                                               "tickets": 0, "overrides": 0, "adm": 0.0})
        d["tickets"] += 1
        if t.overridden:
            d["overrides"] += 1
            d["adm"] += t.adm_amount or 0
    perf = []
    for d in by_agency.values():
        pct = round(100.0 * (1 - d["overrides"] / d["tickets"]), 1) if d["tickets"] else 100.0
        perf.append({**d, "compliance": pct, "light": _traffic_light(pct)})
    perf.sort(key=lambda r: -r["compliance"])

    return render_provider(
        "provider/audits.html", "AUDITS",
        range=range_, filter_agency=filter_agency, sort_health=sort_health,
        agency_names=[a.name for a in agencies],
        rows=rows, reason_dist=REASON_DIST, drill=drill, drilldown_rules=DRILLDOWN_RULES,
        airline_risk=airline_risk, agency_perf=perf,
    )


# --- Routes: escalations -------------------------------------------------

@app.route("/provider/escalations")
def provider_escalations():
    rows = [
        escalation_to_dict(e)
        for e in tenant_q(Escalation).order_by(Escalation.created_at.desc()).all()
    ]
    return render_provider(
        "provider/escalations.html", "ESCALATIONS",
        escalations=rows,
        agency_names=[a.name for a in all_agencies()],
        escalate_id=request.args.get("escalate"),
        escalate_to=request.args.get("to"),
        show_new=request.args.get("new") == "1",
    )


@app.route("/provider/escalations/new", methods=["POST"])
def new_escalation():
    new_id = f"ESC-{random.randint(7000, 7999)}"
    db.session.add(Escalation(
        id=new_id,
        provider_id=current_provider_id(),
        agency=request.form.get("agency") or "All agencies",
        pnr=(request.form.get("pnr") or "—").upper(),
        subject=request.form.get("subject") or "",
        level="L1",
        priority=request.form.get("priority") or "MED",
        opened="just now",
        status="OPEN",
        sla="24 hr left",
    ))
    write_audit("ESCALATION_CREATE", "escalation", new_id)
    db.session.commit()
    return redirect(url_for("provider_escalations"))


@app.route("/provider/escalations/<esc_id>/escalate", methods=["POST"])
def escalate_escalation(esc_id):
    to = request.form.get("to") or "L2"
    needed = "escalation:to_vendor" if to == "VENDOR" else "escalation:to_l2"
    if not can(current_user, needed):
        return render_template("auth/403.html"), 403
    e = get_owned_or_404(Escalation, esc_id)
    e.status = "PENDING"
    e.level = to
    write_audit("ESCALATION_ESCALATE", "escalation", e.id, note=f"→{to}")
    db.session.commit()
    return redirect(url_for("provider_escalations"))


@app.route("/provider/escalations/<esc_id>/resolve", methods=["POST"])
@require("escalation:resolve")
def resolve_escalation(esc_id):
    e = get_owned_or_404(Escalation, esc_id)
    e.status = "RESOLVED"
    write_audit("ESCALATION_RESOLVE", "escalation", e.id)
    db.session.commit()
    return redirect(url_for("provider_escalations"))


# --- Routes: policies, emulate, learning ---------------------------------

@app.route("/provider/policies")
def provider_policies():
    docs = [{"cat": d.cat, "name": d.name, "v": d.v} for d in PolicyDoc.query.all()]
    return render_provider("provider/policies.html", "POLICIES", docs=docs)


@app.route("/provider/emulate")
@require("emulate:use")
def provider_emulate():
    return render_provider("provider/emulate.html", "EMULATE")


@app.route("/provider/audit-log")
@require("user:invite")  # ADMIN-only (same gate as user invite)
def provider_audit_log():
    """Filterable view over the AuditLog table (own tenant only)."""
    action_filter = request.args.get("action", "").strip().upper()
    days = int(request.args.get("days", 7))
    actor_id = request.args.get("actor", "").strip()

    q = tenant_q(AuditLog).order_by(AuditLog.created_at.desc())
    if action_filter:
        q = q.filter(AuditLog.action.ilike(f"%{action_filter}%"))
    if actor_id:
        q = q.filter(AuditLog.actor_user_id == actor_id)
    if days > 0:
        q = q.filter(AuditLog.created_at >= datetime.utcnow() - timedelta(days=days))
    rows = q.limit(500).all()

    actor_map = {u.id: u for u in tenant_q(User).all()}
    actions = [a[0] for a in db.session.query(AuditLog.action).filter_by(
        provider_id=current_provider_id()
    ).distinct().all()]

    return render_provider(
        "provider/audit_log.html", "AUDIT_LOG",
        rows=rows, actors=list(actor_map.values()), actor_map=actor_map,
        actions=sorted(actions), action_filter=action_filter,
        days=days, actor_filter=actor_id,
    )


@app.route("/admin/reset-demo", methods=["POST"])
@require("user:invite")  # ADMIN-only
def admin_reset_demo():
    from seed import seed_all
    seed_all()
    flash("Demo data has been reset to the clean baseline.", "success")
    # After reseed our own user row is regenerated — log out so the session
    # gets re-established cleanly.
    logout_user()
    return redirect(url_for("login"))


@app.route("/admin/inject-event", methods=["POST"])
@require("user:invite")  # ADMIN-only
def admin_inject_event():
    """Fire one stream event on demand — handy mid-pitch."""
    from stream import run_one_tick
    desc = run_one_tick(app)
    flash(f"Event injected: {desc}" if desc else "No active agencies — nothing to inject.", "info")
    return redirect(request.referrer or url_for("provider_dashboard"))


@app.route("/provider/live-feed.json")
def provider_live_feed():
    """JSON feed consumed by the dashboard's Live Activity card.

    Returns the 8 most recent audit-log rows for this tenant. Polled by
    JS every 10s. CSRF not required (read-only, requires session via
    the before_request guard).
    """
    if not current_user.is_authenticated or not current_user.is_provider_staff():
        return jsonify({"events": []}), 403
    rows = (
        tenant_q(AuditLog)
        .order_by(AuditLog.created_at.desc())
        .limit(8)
        .all()
    )
    return jsonify({
        "events": [
            {
                "id": r.id,
                "action": r.action,
                "target": f"{r.target_type}:{r.target_id}",
                "note": r.note or "",
                "when": humanize(r.created_at),
                "iso": r.created_at.isoformat() if r.created_at else None,
                "is_system": r.actor_user_id is None,
            }
            for r in rows
        ],
        "now": datetime.utcnow().isoformat(),
    })


@app.route("/provider/learning")
def provider_learning():
    modules = [{"name": m.name, "progress": m.progress, "badge": m.badge}
               for m in LearningModule.query.all()]
    return render_provider("provider/learning.html", "LEARNING", modules=modules)


# --- Routes: respond -----------------------------------------------------

@app.route("/provider/respond")
def provider_respond():
    threads = tenant_q(Thread).order_by(Thread.created_at.desc()).all()
    if not threads:
        return render_provider("provider/respond.html", "RESPOND", threads=[], active=None)
    active_id = request.args.get("thread", threads[0].id)
    active_obj = next((t for t in threads if t.id == active_id), threads[0])
    return render_provider(
        "provider/respond.html", "RESPOND",
        threads=[thread_to_dict(t) for t in threads],
        active=thread_to_dict(active_obj),
    )


@app.route("/provider/respond/<thread_id>/reply", methods=["POST"])
def reply_thread(thread_id):
    text = request.form.get("text", "").strip()
    if not text:
        return redirect(url_for("provider_respond"))
    t = get_owned_or_404(Thread, thread_id)
    db.session.add(Message(thread_id=t.id, sender="OPS", text=text, t="now"))
    t.last = text
    write_audit("THREAD_REPLY", "thread", t.id)
    db.session.commit()
    return redirect(url_for("provider_respond"))


# --- Health check (for hosts) --------------------------------------------

@app.route("/legal")
def legal():
    """Public legal templates: ToS liability clause, SaaS seat terms,
    data-privacy / tenancy isolation (client updates Batch 5, XC-1)."""
    return render_template("legal.html")


# --- Public JSON API (client updates Batch 5, XC-3) ------------------------
#
# Implements the compliance API exactly as specified in the client's
# document (pp. 3–5): PNR validation, helpdesk escalation, ticket status.
# Auth: X-API-Key header (demo key; per-agency keys come with production).

API_KEY = os.environ.get("AEROGUARD_API_KEY", "demo-key-aeroguard")


def _api_auth_error():
    return jsonify({"status": "ERROR", "error": "Missing or invalid X-API-Key header."}), 401


@app.route("/api/v1/compliance/validate", methods=["POST"])
@csrf.exempt
def api_compliance_validate():
    """POST /api/v1/compliance/validate — evaluate a PNR payload against
    regulatory, airline and corporate-policy rules (spec pp. 3–4)."""
    if request.headers.get("X-API-Key") != API_KEY:
        return _api_auth_error()
    p = request.get_json(silent=True) or {}
    booking = p.get("booking_details") or {}
    itinerary = booking.get("itinerary") or []
    passengers = booking.get("passengers") or []
    elements = booking.get("elements") or {}

    violations = []
    intl = any((s.get("origin") or "")[:2] != (s.get("destination") or "")[:2] for s in itinerary)
    for pax in passengers:
        if not pax.get("passport_number"):
            route = "-".join(filter(None, [itinerary[0].get("origin") if itinerary else None,
                                           itinerary[-1].get("destination") if itinerary else None]))
            violations.append({
                "rule_code": "SEC_PASSPORT_MISSING",
                "severity": "CRITICAL",
                "category": "Security / API Requirement",
                "message": f"Passenger {pax.get('pax_id', '?')} ({pax.get('name', 'UNKNOWN')}) is missing "
                           f"API/Passport data required for international route {route or 'N/A'}.",
                "suggested_action": "Add APIS element using format: SR DOCS <airline> HK1-P-<nat>-<passport>-…",
            })
    if not elements.get("has_email"):
        violations.append({
            "rule_code": "POL_EMAIL_MISSING",
            "severity": "WARNING",
            "category": "Agency Quality Control",
            "message": "Passenger email address (APEMAIL) is missing from the PNR.",
            "suggested_action": "Add contact email element.",
        })
    if not elements.get("has_ticketing_tl"):
        violations.append({
            "rule_code": "TKT_TL_MISSING",
            "severity": "WARNING",
            "category": "Ticketing",
            "message": "No ticketing time limit (TTL) set — carrier auto-cancellation risk.",
            "suggested_action": "Add a TAU/TL element before end of day.",
        })

    criticals = [v for v in violations if v["severity"] == "CRITICAL"]
    warnings_ = [v for v in violations if v["severity"] == "WARNING"]
    # Advisories: match broadcasts whose source or title references one of
    # the itinerary's carriers (codes or common carrier names).
    CARRIER_NAMES = {"ET": "Ethiopian", "EK": "Emirates", "SA": "South African",
                     "LH": "Lufthansa", "QR": "Qatar"}
    carriers = {s.get("airline") for s in itinerary if s.get("airline")}
    needles = {c for c in carriers} | {CARRIER_NAMES[c] for c in carriers if c in CARRIER_NAMES}
    advisories = [
        {"airline": b.source, "notice": b.title}
        for b in Broadcast.query.order_by(Broadcast.created_at.desc()).limit(10).all()
        if any(n in b.source or n in b.title for n in needles)
    ]
    return jsonify({
        "status": "SUCCESS",
        "compliance_status": "FLAGGED" if violations else "COMPLIANT",
        "summary": {
            "total_errors": len(criticals),
            "total_warnings": len(warnings_),
            "block_ticketing": bool(criticals),
        },
        "violations": violations,
        "airline_advisories": advisories,
    })


@app.route("/api/v1/compliance/escalate", methods=["POST"])
@csrf.exempt
def api_compliance_escalate():
    """POST /api/v1/compliance/escalate — waiver/override routed to the
    helpdesk queue (spec p. 5)."""
    if request.headers.get("X-API-Key") != API_KEY:
        return _api_auth_error()
    p = request.get_json(silent=True) or {}
    new_id = f"ESC-{random.randint(8000, 8999)}"
    e = Escalation(
        id=new_id,
        provider_id=DEMO_TERMINAL["provider_id"],
        agency=p.get("agency_id") or DEMO_TERMINAL["agency"],
        pnr=(p.get("pnr_locator") or "—").upper()[:20],
        subject=f"[API] {(p.get('reason_for_escalation') or 'Escalation via API')[:200]}",
        level="L1",
        category="GENERAL",
        priority="HIGH",
        opened="just now",
        status="OPEN",
        sla="4 hr left",
    )
    e.sla_due_at = datetime.utcnow() + timedelta(hours=4)
    db.session.add(e)
    db.session.commit()
    return jsonify({
        "status": "QUEUED",
        "ticket_id": new_id,
        "message": "Escalation logged. Helpdesk supervisor notified. Average review time: 4 minutes.",
        "status_url": f"/api/v1/tickets/{new_id}",
    })


@app.route("/api/v1/tickets/<ticket_id>")
@csrf.exempt
def api_ticket_status(ticket_id):
    if request.headers.get("X-API-Key") != API_KEY:
        return _api_auth_error()
    e = Escalation.query.get(ticket_id)
    if e is None:
        return jsonify({"status": "ERROR", "error": "Unknown ticket id."}), 404
    return jsonify({
        "ticket_id": e.id,
        "status": e.status,
        "priority": e.priority,
        "queue": ESCALATION_TIERS.get(e.category or "GENERAL"),
        "pnr": e.pnr,
        "subject": e.subject,
        "sla_due_at": e.sla_due_at.isoformat() + "Z" if e.sla_due_at else None,
    })


@app.route("/healthz")
def healthz():
    return {"status": "ok"}, 200


# --- CLI commands --------------------------------------------------------

@app.cli.command("seed")
def cli_seed():
    """Populate the database with the standard demo state."""
    from seed import seed_all
    seed_all()
    print("Seeded demo data: 3 providers, 5 users, 6 agencies, 3 escalations.")


@app.cli.command("reset-demo")
def cli_reset_demo():
    """Wipe and re-seed the database in one command."""
    from seed import reset_demo
    reset_demo()
    print("Demo data reset.")


# --- Auto-seed on first boot ---------------------------------------------

def _auto_seed_if_empty():
    """Load the demo data automatically when the database is empty.

    Runs at import time so it also fires under gunicorn on hosted
    deployments (e.g. Render free tier), where there is no shell to run
    ``flask seed`` by hand. Guards:
    1. Skip during CLI commands (``flask db upgrade`` / ``flask seed``).
    2. In dev, only run in the reloader's child process.
    3. Only seed when the users table is empty, so existing data is never
       wiped on a redeploy.
    Any error is logged and swallowed so it can never block startup.
    """
    if os.environ.get("FLASK_RUN_FROM_CLI") == "true":
        return
    if app.config["DEBUG"] and os.environ.get("WERKZEUG_RUN_MAIN") != "true":
        return
    try:
        with app.app_context():
            if User.query.count() == 0:
                from seed import seed_all
                seed_all()
                app.logger.info("Auto-seeded demo data (empty database detected).")
    except Exception as exc:  # never block startup on a seed problem
        app.logger.warning("Auto-seed skipped: %s", exc)


_auto_seed_if_empty()


# --- Background event stream (Section 7) ---------------------------------

def _maybe_start_stream():
    """Start the simulated GDS event stream.

    Two guards apply:
    1. In Flask dev mode the reloader runs the app twice (parent
       watcher + child runner). Only start in the child to avoid two
       schedulers.
    2. Don't start during CLI commands (``flask seed`` etc.) — those
       run with a different invocation path and don't need a ticker.
    """
    if os.environ.get("FLASK_RUN_FROM_CLI") == "true":
        return
    # In dev, the reloader's child sets WERKZEUG_RUN_MAIN=true.
    if app.config["DEBUG"] and os.environ.get("WERKZEUG_RUN_MAIN") != "true":
        return
    from stream import start_event_stream
    start_event_stream(app)


_maybe_start_stream()


if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=int(os.environ.get("PORT", 5050)),
        debug=app.config["DEBUG"],
    )
